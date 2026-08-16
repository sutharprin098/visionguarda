import csv
import json

leads = [
    # 1. CP PLUS (Aditya Infotech Ltd.) - Product Team
    {
        "Company Name": "CP PLUS (Aditya Infotech Ltd.)",
        "Region": "India",
        "Category": "CCTV Hardware OEM",
        "Target Executive Name": "Product Team (Video Analytics)",
        "Target Executive Title": "Sr. Product Manager / Lead Engineering Manager",
        "Direct Email Address": "product.vms@cpplusworld.com",
        "Top Email Syntaxes": "first.last@cpplusworld.com | first.last@adityagroup.com",
        "Target Company Product": "Indigo NVR Series & CP-UNR 4K Enterprise NVRs",
        "Product Tech Gap": "NVRs rely on legacy basic motion detection; lack modern GPU-accelerated YOLOX speed telemetry & zero-cloud desktop VMS studio.",
        "CamAI Plug-in Upgrade Solution": "Integrate CamAI's PyTorch 57 FPS inference engine & Electron desktop studio directly into Indigo NVR client bundle.",
        "Tailored Email Copy": "Hi Product Team, I noticed CP PLUS Indigo NVRs are expanding in enterprise security. We built CamAI - a complete PyTorch/YOLOX Electron desktop & Next.js portal stack with vehicle speed telemetry & ANPR. Instead of spending 10 months building this in-house, you can acquire full 100% source code IP to bundle with Indigo NVRs."
    },
    # 2. CP PLUS (Aditya Infotech Ltd.) - R&D Team
    {
        "Company Name": "CP PLUS (Aditya Infotech Ltd.)",
        "Region": "India",
        "Category": "CCTV Hardware OEM",
        "Target Executive Name": "Monika Sharma",
        "Target Executive Title": "Senior VP – Research & Development",
        "Direct Email Address": "monika.sharma@cpplusworld.com",
        "Top Email Syntaxes": "first.last@cpplusworld.com | first.last@adityagroup.com",
        "Target Company Product": "CP-UNC 4K IP Camera Line & Guard Series",
        "Product Tech Gap": "Requires localized edge AI video analytics without cloud latency or monthly licensing costs.",
        "CamAI Plug-in Upgrade Solution": "100% On-premise zero-cloud egress Python inference backend with multi-camera stream processing.",
        "Tailored Email Copy": "Hi Monika, CP-UNC 4K IP cameras have great optics. CamAI provides a turn-key zero-cloud egress AI video analytics engine (ANPR + Speed overlay + Line crossing) written in Python & Electron. We are looking to sell the complete codebase IP to an R&D team like CP PLUS to save 12 months of development."
    },
    # 3. Matrix Comsec Pvt. Ltd.
    {
        "Company Name": "Matrix Comsec Pvt. Ltd.",
        "Region": "India",
        "Category": "Security OEM",
        "Target Executive Name": "Product Management Team",
        "Target Executive Title": "Product Manager – Video Surveillance (VMS)",
        "Direct Email Address": "pm.surveillance@matrixcomsec.com",
        "Top Email Syntaxes": "first.last@matrixcomsec.com | first_last@matrixcomsec.com",
        "Target Company Product": "SATATYA VMS & SATATYA NVR4208X Series",
        "Product Tech Gap": "SATATYA desktop client lacks interactive live telemetry overlays, zone speed gates, and multi-tenant Supabase backend.",
        "CamAI Plug-in Upgrade Solution": "Plugin CamAI's Electron + React Desktop client and Supabase cloud portal into SATATYA enterprise VMS ecosystem.",
        "Tailored Email Copy": "Hi SATATYA Product Team, SATATYA NVRs are strong hardware. CamAI adds high-speed telemetry studio overlays, speed estimation, and a sleek Next.js control portal. We are offering 100% source code acquisition so your engineering team can integrate it into SATATYA VMS immediately."
    },
    # 4. Videonetics Technology
    {
        "Company Name": "Videonetics Technology",
        "Region": "India",
        "Category": "AI Video VMS Software",
        "Target Executive Name": "Product Lead (Smart City)",
        "Target Executive Title": "Lead Product Manager & AI Architect",
        "Direct Email Address": "product@videonetics.com",
        "Top Email Syntaxes": "first.last@videonetics.com | first@videonetics.com",
        "Target Company Product": "Videonetics Intelligent VMS & TrafficMon",
        "Product Tech Gap": "Needs a lightweight standalone desktop client for security teams needing local GPU auto-detection without thick VMS server installs.",
        "CamAI Plug-in Upgrade Solution": "Use CamAI Electron client + PyTorch auto GPU fallback backend (Intel/NVIDIA/DirectML) for rapid lightweight deployment.",
        "Tailored Email Copy": "Hi Product Team, Videonetics leads smart city video analytics. CamAI is a lightweight, zero-cloud egress Python & Electron desktop VMS with built-in speed telemetry & Supabase cloud sync. We're offering full IP source code transfer for product managers looking to fast-track modern desktop UI & telemetry."
    },
    # 5. Vehant Technologies
    {
        "Company Name": "Vehant Technologies",
        "Region": "India",
        "Category": "Traffic & Physical Security AI OEM",
        "Target Executive Name": "AI Engineering Team",
        "Target Executive Title": "Engineering Manager – Computer Vision & AI",
        "Direct Email Address": "engineering.ai@vehant.com",
        "Top Email Syntaxes": "first@vehant.com | first.last@vehant.com",
        "Target Company Product": "OKEAN AI Platform & TrafficMon Systems",
        "Product Tech Gap": "TrafficMon needs a sleek web telemetry dashboard & edge function cloud alerts for multi-camera nodes.",
        "CamAI Plug-in Upgrade Solution": "Acquire CamAI's Supabase edge functions, database schema migrations, and Next.js portal to power OKEAN AI backend.",
        "Tailored Email Copy": "Hi Engineering Team, Vehant's TrafficMon is impressive for traffic management. CamAI contains a full Supabase DB schema, Next.js telemetry portal, and PyTorch speed calibration pipeline. We're selling the complete IP to help your team skip months of UI and backend dev."
    },
    # 6. Sparsh Surveillance
    {
        "Company Name": "Sparsh Surveillance",
        "Region": "India",
        "Category": "Make-in-India CCTV OEM",
        "Target Executive Name": "Product Engineering Lead",
        "Target Executive Title": "Head of Product & Software Engineering",
        "Direct Email Address": "software.lead@sparshsecuritech.com",
        "Top Email Syntaxes": "first.last@sparshsecuritech.com | first.last@sparshcctv.com",
        "Target Company Product": "Sparsh Freedom & Enterprise NVR Series",
        "Product Tech Gap": "Government tenders require 'Make-in-India' 100% indigenous software stack with zero foreign cloud dependencies.",
        "CamAI Plug-in Upgrade Solution": "CamAI is 100% locally developed in India with on-premise zero-cloud egress architecture, perfect for Indian Govt tenders.",
        "Tailored Email Copy": "Hi Product Team, Sparsh is leading Make-in-India hardware. CamAI is a 100% indigenously developed Python + Electron + Next.js video analytics codebase designed for zero-cloud egress. Buying this codebase IP gives Sparsh a ready-to-bid software stack for Govt Smart City tenders."
    },
    # 7. Godrej Security Solutions
    {
        "Company Name": "Godrej Security Solutions",
        "Region": "India",
        "Category": "Security Solutions",
        "Target Executive Name": "Category Product Manager",
        "Target Executive Title": "Product Manager – Electronic Security & CCTV",
        "Direct Email Address": "gss.product@godrej.com",
        "Top Email Syntaxes": "first.last@godrej.com | firstinitiallast@godrej.com",
        "Target Company Product": "Godrej Eve & Commercial Enterprise CCTV Line",
        "Product Tech Gap": "Godrej CCTV hardware relies on standard third-party VMS; lacks Godrej-branded desktop client & telemetry studio.",
        "CamAI Plug-in Upgrade Solution": "White-label CamAI Electron desktop app and Next.js portal with Godrej branding for enterprise CCTV bundles.",
        "Tailored Email Copy": "Hi Godrej Product Team, Godrej CCTV hardware is trusted across India. We built CamAI - a complete desktop & web software ecosystem for CCTV monitoring, ANPR, and speed telemetry. You can acquire full codebase IP to launch a custom white-labeled Godrej VMS software platform."
    },
    # 8. Honeywell Building Technologies India
    {
        "Company Name": "Honeywell Building Technologies India",
        "Region": "India",
        "Category": "Global Security Systems OEM",
        "Target Executive Name": "Sr. Product Manager (Commercial Security)",
        "Target Executive Title": "Sr. Product Manager – Video Surveillance Solutions",
        "Direct Email Address": "honeywell.security@honeywell.com",
        "Top Email Syntaxes": "first.last@honeywell.com | first.middle.last@honeywell.com",
        "Target Company Product": "Honeywell MAXPRO NVR & Performance Series IP Cameras",
        "Product Tech Gap": "MAXPRO client setup is complex; lacks lightweight cross-platform Electron studio with automated GPU backend fallback.",
        "CamAI Plug-in Upgrade Solution": "CamAI provides an auto-detecting PyTorch engine (Intel OpenVINO/DirectML/CUDA) + lightweight Electron desktop VMS.",
        "Tailored Email Copy": "Hi Honeywell Product Team, MAXPRO NVRs are industrial benchmarks. CamAI is a zero-cloud egress Python & Electron video analytics stack featuring auto GPU fallback, ANPR, and vehicle speed telemetry. We are transferring 100% codebase ownership for teams seeking a fast-track VMS software upgrade."
    },
    # 9. L&T Technology Services (LTTS Smart World)
    {
        "Company Name": "L&T Technology Services (LTTS Smart World)",
        "Region": "India",
        "Category": "Smart City System Integrator",
        "Target Executive Name": "Head of Product (Smart Cities)",
        "Target Executive Title": "Lead Product Architect – ICCC Smart Governance",
        "Direct Email Address": "smartcities@lnttechservices.com",
        "Top Email Syntaxes": "first.last@lnttechservices.com | first.last@lnt.com",
        "Target Company Product": "Municipal ICCC (Integrated Command and Control Center) Software Nodes",
        "Product Tech Gap": "ICCC command centers require local edge analytics nodes for speed violation detection and multi-camera stream health check.",
        "CamAI Plug-in Upgrade Solution": "CamAI Edge PyTorch engine & Supabase real-time telemetry acts as a ready command node for ICCC Smart City deployments.",
        "Tailored Email Copy": "Hi LTTS Smart City Team, ICCC command centers require high-speed local stream analytics. CamAI is a complete zero-cloud egress Python + Next.js + Electron stack built for real-time speed estimation, camera health heartbeats, and alert logging. Full IP buyout is available for immediate ICCC integration."
    },
    # 10. Staqu Technologies
    {
        "Company Name": "Staqu Technologies",
        "Region": "India",
        "Category": "AI Video Analytics Software",
        "Target Executive Name": "Product Lead (AI Platforms)",
        "Target Executive Title": "Product Manager – Video AI & Security Platforms",
        "Direct Email Address": "product.ai@staqu.com",
        "Top Email Syntaxes": "first.last@staqu.com | first@staqu.com",
        "Target Company Product": "JARVIS AI Video Analytics Platform",
        "Product Tech Gap": "JARVIS is cloud-heavy; enterprise clients demand a high-performance offline desktop client with local live canvas drawing tools.",
        "CamAI Plug-in Upgrade Solution": "Acquire CamAI's Electron + React Desktop client and local OpenCV/YOLO pipeline to serve offline enterprise clients.",
        "Tailored Email Copy": "Hi Staqu Product Team, Staqu's JARVIS AI is great for cloud video analytics. CamAI offers an offline, zero-cloud egress Electron desktop VMS studio with interactive canvas drawing, ANPR, and speed telemetry. Acquiring our full codebase IP allows Staqu to instantly offer an offline desktop VMS client to enterprise buyers."
    }
]

# Write CSV
csv_file = "d:/camAI/docs/CamAI_Enterprise_Outreach_Leads.csv"
fieldnames = list(leads[0].keys())

with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(leads)

print(f"CSV successfully generated at: {csv_file}")

# Try creating XLSX if openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CamAI 10 Product Manager Leads"

    # Header styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.append(fieldnames)
    for col_num in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, lead in enumerate(leads, start=2):
        row_data = [lead[col] for col in fieldnames]
        ws.append(row_data)
        for col_num in range(1, len(fieldnames) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num in [1, 2, 4]:
                cell.alignment = Alignment(vertical="top")
                if col_num in [1, 4]:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set column widths
    widths = [26, 12, 22, 24, 28, 28, 30, 30, 35, 35, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    xlsx_file = "d:/camAI/docs/CamAI_Enterprise_Outreach_Leads.xlsx"
    wb.save(xlsx_file)
    print(f"XLSX successfully generated at: {xlsx_file}")

except Exception as e:
    print(f"XLSX creation skipped: {e}")
